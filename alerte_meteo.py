import requests
import json
import time
import urllib3
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from functools import lru_cache
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

OAUTH2_TOKEN   = ""
API_KEY        = ""

SEUIL_ANOMALIE   = 10     # °C — seuil de base pour alertes beau/mauvais temps
MAX_WORKERS      = 10    # Requêtes parallèles températures
MAX_WORKERS_NORM = 1     # 1 seul worker ERA5 pour éviter les 429

PERIODE_NORMALE = "2020-2025"

PERIODES = {
    "1991-2020": ("1991-01-01", "2020-12-31"),
    "1991-2025": ("1991-01-01", "2025-12-31"),
    "2016-2025": ("2016-01-01", "2025-12-31"),
    "2020-2025": ("2020-01-01", "2025-12-31"),
}

# Résolution ERA5-Land : 0.1° ≈ 9 km
ERA5_RESOLUTION = 1  # 1 décimale = 0.1°

# Communes insulaires — coordonnées exactes (pas d'arrondi)
# car l'arrondi à 0.1° ferait tomber la maille en mer
COMMUNES_INSULAIRES = {
    "Bangor",
    "Groix",
    "Hœdic",
    "Le Palais",
    "Locmaria",
    "Sauzon",
    "Île-d'Houat",
    "Île-aux-Moines",
    "Île-d'Arz",
}

# Session persistante ERA5 — verify=False pour contourner les coupures SSL
_era5_session        = requests.Session()
_era5_session.verify = False

# ─────────────────────────────────────────────
# COMMUNES DU MORBIHAN (56)
# ─────────────────────────────────────────────

def charger_communes_morbihan():
    url    = "https://geo.api.gouv.fr/departements/56/communes"
    params = {"fields": "nom,centre", "format": "json", "geometry": "centre"}
    try:
        r = requests.get(url, params=params, timeout=20)
        r.raise_for_status()
        communes = {}
        for commune in r.json():
            centre = commune.get("centre")
            if not centre:
                continue
            coords = centre.get("coordinates")
            if not coords:
                continue
            lon, lat = coords
            communes[commune["nom"]] = (lat, lon)
        print(f"✅ {len(communes)} communes chargées automatiquement")
        return communes
    except Exception as e:
        print(f"  ❌ Erreur chargement communes : {e}")
        return {}

print("Chargement des communes du Morbihan...")
COMMUNES_56 = charger_communes_morbihan()

DEPTS_BRETAGNE     = ["22", "29", "35", "56"]
NOMS_DEPTS         = {"22": "Côtes-d'Armor", "29": "Finistère", "35": "Ille-et-Vilaine", "56": "Morbihan"}
COULEURS_VIGILANCE = {1: "Vert", 2: "Jaune", 3: "Orange", 4: "Rouge"}
PHENOMENES         = {
    1: "Vent violent", 2: "Pluie-inondation", 3: "Orages", 4: "Crues",
    5: "Neige-verglas", 6: "Canicule", 7: "Grand froid", 8: "Avalanches",
    9: "Vagues-submersion", 10: "Pluie-inondation (crues)",
}


def get_headers():
    if OAUTH2_TOKEN:
        return {"Authorization": f"Bearer {OAUTH2_TOKEN}", "accept": "application/json"}
    elif API_KEY:
        return {"apikey": API_KEY, "accept": "application/json"}
    else:
        raise ValueError("Configurez OAUTH2_TOKEN ou API_KEY dans le script !")


# ─────────────────────────────────────────────
# 1. VIGILANCE
# ─────────────────────────────────────────────

def get_vigilance():
    url = "https://public-api.meteofrance.fr/public/DPVigilance/v1/cartevigilance/encours"
    try:
        resp = requests.get(url, headers=get_headers(), timeout=10)
        resp.raise_for_status()
        return resp.json()
    except ValueError as e:
        print(f"  [ERREUR] {e}")
    except requests.exceptions.HTTPError:
        print(f"  [ERREUR] HTTP {resp.status_code} — token invalide ou expiré")
    except requests.exceptions.RequestException as e:
        print(f"  [ERREUR] {e}")
    return None


def extraire_alertes_bretagne(data):
    alertes = {dept: [] for dept in DEPTS_BRETAGNE}
    if not data:
        return alertes

    def chercher_domaines(obj):
        if isinstance(obj, list):
            if obj and isinstance(obj[0], dict) and "id" in obj[0]:
                return obj
            for item in obj:
                r = chercher_domaines(item)
                if r:
                    return r
        elif isinstance(obj, dict):
            for key in ["domain_ids", "domaines", "massifs", "items"]:
                if key in obj:
                    return obj[key]
            for val in obj.values():
                r = chercher_domaines(val)
                if r:
                    return r
        return []

    for domaine in chercher_domaines(data):
        dept_id = str(domaine.get("id", "")).zfill(2)
        if dept_id not in DEPTS_BRETAGNE:
            continue
        for ech in domaine.get("echeances", []):
            niveau = ech.get("color_id", 1)
            if niveau >= 3:
                pid = ech.get("phenomenon_id")
                alertes[dept_id].append({
                    "niveau":     COULEURS_VIGILANCE.get(niveau, f"Niveau {niveau}"),
                    "couleur_id": niveau,
                    "phenomene":  PHENOMENES.get(pid, f"Phénomène {pid}"),
                    "debut":      ech.get("begin_time", "?"),
                    "fin":        ech.get("end_time", "?"),
                })
    return alertes


# ─────────────────────────────────────────────
# 2. TEMPÉRATURE ACTUELLE
# ─────────────────────────────────────────────

def get_temperature_actuelle(lat, lon):
    """Récupère la température actuelle en temps réel via Open-Meteo forecast."""
    try:
        r = requests.get("https://api.open-meteo.com/v1/forecast", params={
            "latitude":      lat,
            "longitude":     lon,
            "current":       "temperature_2m",
            "forecast_days": 1,
            "timezone":      "Europe/Paris",
        }, timeout=10)
        r.raise_for_status()
        return r.json().get("current", {}).get("temperature_2m")
    except Exception:
        return None


# ─────────────────────────────────────────────
# 3. NORMALES ERA5-LAND HORAIRES (1 requête/an)
# ─────────────────────────────────────────────
# Stratégie : pour chaque année de la période, on ne demande que
# ±7 jours autour du jour J → ~15 jours × 24h = 360 valeurs max.
# On filtre ensuite l'heure exacte du jour exact.
# Résultat : requêtes ultra-légères, zéro timeout, zéro 504.

def arrondir_coords(lat, lon, commune=""):
    """
    Arrondit les coordonnées à ERA5_RESOLUTION décimales (0.1° ≈ 9 km).
    EXCEPTION : communes insulaires → coordonnées exactes pour éviter
    que l'arrondi tombe en mer (pas de données ERA5-Land en mer).
    """
    if commune in COMMUNES_INSULAIRES:
        return (lat, lon)
    return (
        round(lat, ERA5_RESOLUTION),
        round(lon, ERA5_RESOLUTION),
    )


def _requete_era5_annee(lat, lon, annee, month, day, hour):
    """
    Requête ERA5-Land ultra-légère : ±7 jours autour du jour J
    pour une année donnée. ~360 valeurs max par requête.
    Filtre sur l'heure exacte du jour exact.
    Retry x4 avec backoff exponentiel + pause 0.5s après succès.
    """
    # Fenêtre ±7 jours — évite le 29 fév pour les années non bissextiles
    try:
        centre = date(annee, month, day)
    except ValueError:
        # 29 fév sur année non bissextile → on prend le 28 fév
        centre = date(annee, 2, 28)

    debut = (centre - timedelta(days=7)).strftime("%Y-%m-%d")
    fin   = (centre + timedelta(days=7)).strftime("%Y-%m-%d")

    for attempt in range(4):
        try:
            r = _era5_session.get(
                "https://archive-api.open-meteo.com/v1/archive",
                params={
                    "latitude":   lat,
                    "longitude":  lon,
                    "start_date": debut,
                    "end_date":   fin,
                    "hourly":     "temperature_2m",
                    "timezone":   "Europe/Paris",
                    "models":     "era5_land",
                },
                timeout=20,
            )

            if r.status_code == 429:
                wait = 2 ** attempt
                print(f"    ⏳ Rate limit ERA5, attente {wait}s...")
                time.sleep(wait)
                continue

            r.raise_for_status()
            data  = r.json().get("hourly", {})
            dates = data.get("time", [])
            vals  = data.get("temperature_2m", [])

            # Filtrage exact : même mois + même jour + même heure
            # Format timestamp : "2020-05-23T14:00"
            valeurs = [
                v for d, v in zip(dates, vals)
                if v is not None
                and int(d[5:7])  == month
                and int(d[8:10]) == day
                and int(d[11:13]) == hour
            ]

            # Cas du 29 fév sur année non bissextile :
            # on prend 28 fév à la même heure comme approximation
            if not valeurs and month == 2 and day == 29:
                valeurs = [
                    v for d, v in zip(dates, vals)
                    if v is not None
                    and int(d[5:7]) == 2 and int(d[8:10]) == 28
                    and int(d[11:13]) == hour
                ]

            time.sleep(0.5)  # pause pour ménager l'API
            return valeurs

        except Exception as e:
            wait = 2 ** attempt
            print(f"    ⚠️  ERA5 {annee} tentative {attempt + 1}/4 "
                  f"({lat:.4f},{lon:.4f}): {e}")
            time.sleep(wait)

    return []


@lru_cache(maxsize=512)
def get_normale_era5_zone(lat_r, lon_r, periode="1991-2020"):
    """
    Calcule la normale ERA5-Land HORAIRE pour la zone (lat_r, lon_r).

    Exemple : si on est le 23 mai à 14h, on calcule la moyenne des
    températures ERA5 à 14h le 23 mai sur toutes les années 1991-2020.

    Stratégie : 1 requête légère (±7 jours) par année → zéro timeout.
    Résultat mis en cache — partagé entre toutes les communes de la zone.

    Temps estimé par période :
      "2020-2025" →  6 req/zone → ~3 min total
      "2016-2025" → 10 req/zone → ~5 min total
      "1991-2020" → 30 req/zone → ~26 min total
      "1991-2025" → 35 req/zone → ~30 min total
    """
    now   = datetime.now()
    month = now.month
    day   = now.day
    hour  = now.hour

    if periode not in PERIODES:
        periode = "1991-2020"

    start_year = int(PERIODES[periode][0][:4])
    end_year   = int(PERIODES[periode][1][:4])

    toutes_valeurs = []
    for annee in range(start_year, end_year + 1):
        valeurs = _requete_era5_annee(lat_r, lon_r, annee, month, day, hour)
        toutes_valeurs.extend(valeurs)

    return round(sum(toutes_valeurs) / len(toutes_valeurs), 1) if toutes_valeurs else None


def get_normale_commune(commune, lat, lon):
    """
    Retourne la normale ERA5 horaire pour une commune.
    Arrondit les coordonnées sauf pour les îles.
    """
    lat_r, lon_r = arrondir_coords(lat, lon, commune)
    return get_normale_era5_zone(lat_r, lon_r, PERIODE_NORMALE)


# ─────────────────────────────────────────────
# 4. COLLECTE PAR COMMUNE
# ─────────────────────────────────────────────

def get_meteo_commune(args):
    commune, (lat, lon) = args
    temp     = get_temperature_actuelle(lat, lon)
    normale  = get_normale_commune(commune, lat, lon)
    anomalie = round(temp - normale, 1) if (temp is not None and normale is not None) else None
    return commune, {
        "lat":                  lat,
        "lon":                  lon,
        "temperature_actuelle": temp,
        "normale_saison":       normale,
        "anomalie":             anomalie,
        "ile":                  commune in COMMUNES_INSULAIRES,
    }


# ─────────────────────────────────────────────
# 5. CLASSIFICATION DES ANOMALIES
# ─────────────────────────────────────────────

def classifier_anomalie(anomalie):
    if anomalie is None:
        return "D9D9D9", "N/A"
    if anomalie >= 5:
        return "FCE4D6", "ALERTE BEAU TEMPS"
    if anomalie >= 3:
        return "FFF2CC", "Hausse notable"
    if anomalie <= -5:
        return "2F5D8A", "ALERTE MAUVAIS TEMPS"
    if anomalie <= -3:
        return "A7C7E7", "En dessous normale"
    return "E2EFDA", "Normal"


# ─────────────────────────────────────────────
# 6. EXPORT EXCEL
# ─────────────────────────────────────────────

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def style_header(cell, bg="1F4E79", fg="FFFFFF"):
    cell.font      = Font(name="Arial", bold=True, color=fg, size=11)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = thin_border()

def exporter_excel(now_str, alertes_par_dept, resultats, nb_zones, heure_normale):
    wb        = Workbook()

    # ── Onglet Températures ───────────────────────────────────────
    ws1       = wb.active
    ws1.title = "Températures Morbihan"

    ws1.merge_cells("A1:F1")
    ws1["A1"]           = "MÉTÉO MORBIHAN (56) — Températures & Anomalies"
    ws1["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:F2")
    ws1["A2"] = (
        f"Collecte du {now_str}  |  "
        f"Normales ERA5-Land horaires à {heure_normale}h — {PERIODE_NORMALE}  |  "
        f"Résolution ~9 km  |  Îles : coordonnées exactes  |  "
        f"Alerte beau temps ≥ +5°C  |  Alerte mauvais temps ≤ -5°C"
    )
    ws1["A2"].font      = Font(name="Arial", italic=True, size=10, color="595959")
    ws1["A2"].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 18

    for col, h in enumerate([
        "Commune",
        "Temp. actuelle (°C)",
        f"Normale {heure_normale}h (°C)",
        "Anomalie (°C)",
        "Statut"
    ], 1):
        style_header(ws1.cell(row=4, column=col, value=h))
    ws1.row_dimensions[4].height = 22

    row = 5
    for commune, info in sorted(resultats.items()):
        temp     = info.get("temperature_actuelle")
        normale  = info.get("normale_saison")
        anomalie = info.get("anomalie")
        est_ile  = info.get("ile", False)
        bg, statut = classifier_anomalie(anomalie)

        nom_affiche = f"🏝 {commune}" if est_ile else commune

        for col, val in enumerate([nom_affiche, temp, normale, anomalie, statut], 1):
            c           = ws1.cell(row=row, column=col, value=val)
            c.font      = Font(name="Arial", size=10)
            c.fill      = PatternFill("solid", start_color=bg)
            c.border    = thin_border()
            c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            if col in (2, 3, 4) and val is not None:
                c.number_format = '0.0"°C"'
        ws1.row_dimensions[row].height = 18
        row += 1

    # Ligne résumé
    row += 1
    alertes_beau    = sum(1 for i in resultats.values() if i.get("anomalie") is not None and i["anomalie"] >= 5)
    alertes_mauvais = sum(1 for i in resultats.values() if i.get("anomalie") is not None and i["anomalie"] <= -5)
    sans_normale    = sum(1 for i in resultats.values() if i.get("normale_saison") is None)
    ws1.merge_cells(f"A{row}:F{row}")
    c       = ws1[f"A{row}"]
    c.value = (
        f"Total communes : {len(resultats)}  |  "
        f"Zones ERA5 : {nb_zones}  |  "
        f"Normales à {heure_normale}h ({PERIODE_NORMALE})  |  "
        f"Alertes beau temps (≥+5°C) : {alertes_beau}  |  "
        f"Alertes mauvais temps (≤-5°C) : {alertes_mauvais}  |  "
        f"Sans normale : {sans_normale}"
    )
    c.font      = Font(name="Arial", bold=True, size=11, color="1F4E79")
    c.alignment = Alignment(horizontal="center")

    # Légende
    row += 2
    ws1.cell(row=row, column=1, value="Légende :").font = Font(name="Arial", bold=True, size=10)
    for i, (label, color, desc) in enumerate([
        ("Orange",     "FCE4D6", "Anomalie ≥ +5°C — Alerte beau temps"),
        ("Jaune",      "FFF2CC", "Anomalie ≥ +3°C — Hausse notable"),
        ("Vert",       "E2EFDA", "Anomalie entre -3°C et +3°C — Normal"),
        ("Bleu Clair", "A7C7E7", "Anomalie ≤ -3°C — En dessous normale"),
        ("Bleu Foncé", "2F5D8A", "Anomalie ≤ -5°C — Alerte mauvais temps"),
        ("Gris",       "D9D9D9", "Normale indisponible (N/A)"),
    ], 1):
        r         = row + i
        c1        = ws1.cell(row=r, column=1, value=label)
        c1.fill   = PatternFill("solid", start_color=color)
        c1.font   = Font(name="Arial", size=10, bold=True)
        c1.border = thin_border()
        ws1.cell(row=r, column=2, value=desc).font = Font(name="Arial", size=10)

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 20
    ws1.column_dimensions["D"].width = 18
    ws1.column_dimensions["E"].width = 26

    # ── Onglet Vigilance ─────────────────────────────────────────
    ws2       = wb.create_sheet("Vigilance")
    ws2.merge_cells("A1:D1")
    ws2["A1"]           = f"ALERTES VIGILANCE MÉTÉO-FRANCE — {now_str}"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:D2")
    ws2["A2"]           = "Seules les alertes orange et rouge sont affichées. Absence = bonne vigilance."
    ws2["A2"].font      = Font(name="Arial", italic=True, size=10, color="595959")
    ws2["A2"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[2].height = 18

    for col, h in enumerate(["Département", "Phénomène", "Niveau", "Période"], 1):
        style_header(ws2.cell(row=4, column=col, value=h))
    ws2.row_dimensions[4].height = 22

    row2          = 5
    a_des_alertes = False
    for dept, alertes in alertes_par_dept.items():
        for a in alertes:
            a_des_alertes = True
            bg = "FCE4D6" if a["couleur_id"] == 3 else "FF9999"
            for col, val in enumerate([
                f"{dept} — {NOMS_DEPTS[dept]}", a["phenomene"], a["niveau"],
                f"{a['debut']} → {a['fin']}"
            ], 1):
                c           = ws2.cell(row=row2, column=col, value=val)
                c.font      = Font(name="Arial", size=10)
                c.fill      = PatternFill("solid", start_color=bg)
                c.border    = thin_border()
                c.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
            ws2.row_dimensions[row2].height = 18
            row2 += 1

    if not a_des_alertes:
        ws2.merge_cells(f"A{row2}:D{row2}")
        c           = ws2[f"A{row2}"]
        c.value     = "✅ Aucune alerte orange ou rouge en Bretagne."
        c.font      = Font(name="Arial", size=11, color="375623")
        c.fill      = PatternFill("solid", start_color="E2EFDA")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = thin_border()
        ws2.row_dimensions[row2].height = 22

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 36

    filename = f"meteo_morbihan_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"\n  📊 Excel exporté : {filename}")
    return filename


# ─────────────────────────────────────────────
# 7. AFFICHAGE TERMINAL
# ─────────────────────────────────────────────

def afficher_alertes(alertes_par_dept):
    print("\n🚨 ALERTES VIGILANCE MÉTÉO-FRANCE (Orange & Rouge)")
    print("─" * 60)
    for dept, alertes in alertes_par_dept.items():
        if alertes:
            print(f"\n  Département {dept} — {NOMS_DEPTS[dept]}")
            for a in alertes:
                emoji = "🟠" if a["couleur_id"] == 3 else "🔴"
                print(f"    {emoji} {a['niveau']} — {a['phenomene']}")
                print(f"       Du {a['debut']} au {a['fin']}")
        else:
            print(f"  ✅ {dept} ({NOMS_DEPTS[dept]}) — Aucune alerte")


def afficher_resume(resultats, heure_normale):
    alertes_beau = [
        (v, i) for v, i in resultats.items()
        if i.get("anomalie") is not None and i["anomalie"] >= SEUIL_ANOMALIE
    ]
    alertes_mauvais = [
        (v, i) for v, i in resultats.items()
        if i.get("anomalie") is not None and i["anomalie"] <= -SEUIL_ANOMALIE
    ]

    print(f"\n\n☀️  ALERTES BEAU TEMPS — anomalie ≥ +{SEUIL_ANOMALIE}°C ({len(alertes_beau)} commune(s))")
    print("─" * 60)
    if alertes_beau:
        for commune, info in sorted(alertes_beau, key=lambda x: -x[1]["anomalie"]):
            ile = " 🏝" if info.get("ile") else ""
            print(f"  ☀️  {commune:<35}{ile} +{info['anomalie']}°C vs normale {heure_normale}h {PERIODE_NORMALE}")
    else:
        print(f"  Aucune commune avec anomalie ≥ +{SEUIL_ANOMALIE}°C.")

    print(f"\n☔️  ALERTES MAUVAIS TEMPS — anomalie ≤ -{SEUIL_ANOMALIE}°C ({len(alertes_mauvais)} commune(s))")
    print("─" * 60)
    if alertes_mauvais:
        for commune, info in sorted(alertes_mauvais, key=lambda x: x[1]["anomalie"]):
            signe = "+" if info["anomalie"] > 0 else ""
            ile   = " 🏝" if info.get("ile") else ""
            print(f"  ⚠️  {commune:<35}{ile} {signe}{info['anomalie']}°C vs normale {heure_normale}h {PERIODE_NORMALE}")
    else:
        print(f"  Aucune commune avec anomalie ≤ -{SEUIL_ANOMALIE}°C.")

    normales = [
        (v, i) for v, i in resultats.items()
        if i.get("anomalie") is not None
        and -SEUIL_ANOMALIE < i["anomalie"] < SEUIL_ANOMALIE
    ]
    sans_normale = [v for v, i in resultats.items() if i.get("normale_saison") is None]

    print(f"\n🟢 COMMUNES NORMALES — anomalie entre -{SEUIL_ANOMALIE}°C et +{SEUIL_ANOMALIE}°C ({len(normales)} commune(s))")

    if sans_normale:
        print(f"\n⚪ SANS NORMALE ({len(sans_normale)} commune(s)) — ERA5 indisponible")
        for c in sorted(sans_normale):
            ile = " 🏝" if c in COMMUNES_INSULAIRES else ""
            print(f"  ⚪ {c}{ile}")


# ─────────────────────────────────────────────
# 8. MAIN
# ─────────────────────────────────────────────

def main():
    now           = datetime.now()
    now_str       = now.strftime("%d/%m/%Y %H:%M")
    heure_normale = now.hour

    start_year = int(PERIODES[PERIODE_NORMALE][0][:4])
    end_year   = int(PERIODES[PERIODE_NORMALE][1][:4])
    nb_annees  = end_year - start_year + 1

    print("=" * 60)
    print(f"  MÉTÉO MORBIHAN (56) — {now_str}")
    print(f"  {len(COMMUNES_56)} communes — requêtes parallèles ({MAX_WORKERS} workers)")
    print(f"  Normales ERA5-Land HORAIRES à {heure_normale}h — période {PERIODE_NORMALE}")
    print(f"  Stratégie : 1 req légère/an (±7j) × {nb_annees} ans")
    print(f"  Îles : coordonnées exactes")
    print("=" * 60)

    # ── Vigilance ─────────────────────────────────────────────────
    print("\n⏳ Récupération des alertes vigilance...")
    if not OAUTH2_TOKEN and not API_KEY:
        print("  ⚠️  Aucun token configuré — vigilance ignorée")
        alertes_par_dept = {dept: [] for dept in DEPTS_BRETAGNE}
    else:
        print(f"  🔑 Mode : {'OAuth2' if OAUTH2_TOKEN else 'API Key'}")
        alertes_par_dept = extraire_alertes_bretagne(get_vigilance())
    afficher_alertes(alertes_par_dept)

    # ── Pré-calcul des normales ERA5-Land horaires ────────────────
    zones_uniques = list({
        arrondir_coords(lat, lon, commune)
        for commune, (lat, lon) in COMMUNES_56.items()
    })
    nb_iles     = sum(1 for c in COMMUNES_56 if c in COMMUNES_INSULAIRES)
    nb_requetes = len(zones_uniques) * nb_annees
    duree_min   = round(nb_requetes * 0.5 / 60, 1)

    print(f"\n⏳ Pré-calcul des normales ERA5-Land horaires {PERIODE_NORMALE} à {heure_normale}h...")
    print(f"   {len(COMMUNES_56)} communes → {len(zones_uniques)} zones ERA5 uniques")
    print(f"   dont {nb_iles} île(s) avec coordonnées exactes")
    print(f"   {nb_annees} ans × {len(zones_uniques)} zones = {nb_requetes} requêtes")
    print(f"   Durée estimée : ~{duree_min} min | 1 worker | pause 0.5s | timeout 20s")

    done_norm = 0
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_NORM) as executor:
        futures_norm = {
            executor.submit(get_normale_era5_zone, lat_r, lon_r, PERIODE_NORMALE): (lat_r, lon_r)
            for lat_r, lon_r in zones_uniques
        }
        for future in as_completed(futures_norm):
            done_norm += 1
            if done_norm % 10 == 0 or done_norm == len(zones_uniques):
                print(f"   [{done_norm:>3}/{len(zones_uniques)}] zones calculées...")

    ok_count = sum(
        1 for lat_r, lon_r in zones_uniques
        if get_normale_era5_zone(lat_r, lon_r, PERIODE_NORMALE) is not None
    )
    print(f"   ✅ {ok_count}/{len(zones_uniques)} zones ERA5 disponibles\n")

    if ok_count < len(zones_uniques):
        manquantes = len(zones_uniques) - ok_count
        print(f"   ⚠️  {manquantes} zones manquantes.")
        print(f"       Fallback rapide → changez PERIODE_NORMALE = \"2020-2025\"\n")

    # ── Températures actuelles en parallèle ───────────────────────
    print(f"⏳ Récupération des températures ({len(COMMUNES_56)} communes en parallèle)...")
    resultats = {}
    done      = 0

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(get_meteo_commune, item): item[0]
            for item in COMMUNES_56.items()
        }
        for future in as_completed(futures):
            commune, data = future.result()
            resultats[commune] = data
            done += 1

            temp     = data["temperature_actuelle"]
            normale  = data["normale_saison"]
            anomalie = data["anomalie"]
            ile_tag  = " 🏝" if data.get("ile") else ""

            if temp is not None:
                signe    = "+" if (anomalie or 0) > 0 else ""
                norm_str = f"{normale}°C"         if normale  is not None else "N/A"
                anom_str = f"{signe}{anomalie}°C" if anomalie is not None else "N/A"
                print(f"  [{done:>3}/{len(COMMUNES_56)}] {commune:<30}{ile_tag} "
                      f"{temp}°C  (normale {heure_normale}h: {norm_str}, anomalie: {anom_str})")
            else:
                print(f"  [{done:>3}/{len(COMMUNES_56)}] {commune}{ile_tag} — N/A")

    afficher_resume(resultats, heure_normale)

    print("\n" + "=" * 60)
    print("  Collecte terminée.")
    print("=" * 60)

    # ── Exports ───────────────────────────────────────────────────
    exporter_excel(now_str, alertes_par_dept, resultats, len(zones_uniques), heure_normale)

    filename_json = f"meteo_morbihan_{now.strftime('%Y%m%d_%H%M')}.json"
    with open(filename_json, "w", encoding="utf-8") as f:
        json.dump({
            "date_collecte":       now_str,
            "heure_normale":       heure_normale,
            "periode_normale":     PERIODE_NORMALE,
            "nb_annees":           nb_annees,
            "source_normale":      f"ERA5-Land horaire ±7j/an via archive-api.open-meteo.com (~9 km) — {len(zones_uniques)} zones",
            "communes_insulaires": list(sorted(COMMUNES_INSULAIRES)),
            "alertes_vigilance":   alertes_par_dept,
            "temperatures": {
                v: {k: val for k, val in i.items() if k not in ("lat", "lon")}
                for v, i in resultats.items()
            }
        }, f, ensure_ascii=False, indent=2)
    print(f"  📁 JSON exporté   : {filename_json}")


if __name__ == "__main__":
    main()